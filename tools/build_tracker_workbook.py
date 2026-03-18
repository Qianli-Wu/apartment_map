#!/usr/bin/env python3
import json
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if "/tmp/codexdeps" not in sys.path:
    sys.path.insert(0, "/tmp/codexdeps")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DATA_PATH = ROOT / "data" / "initial_tracker_data.json"
OUTPUT_PATH = ROOT / "output" / "spreadsheet" / "South_Bay_2B2B_Apartment_Search.xlsx"


CONTROL_HEADERS = ["Setting", "Value", "Description"]
SEED_HEADERS = [
    "Source Type",
    "Source File/URL",
    "Property Name",
    "City",
    "Address",
    "Claimed Rent",
    "Claimed Promo",
    "Claimed Transit Note",
    "Claimed Review Note",
    "Needs Verification",
]
QUEUE_HEADERS = [
    "Priority",
    "Property Name",
    "City",
    "Source Type",
    "Search Task",
    "Official URL Found",
    "Transit Check Done",
    "Review Check Done",
    "Owner",
    "Queue Status",
    "Notes",
]
INTAKE_HEADERS = [
    "Research Date",
    "Researcher",
    "Property Type",
    "Apartment Name",
    "City",
    "Address",
    "Unit/Floorplan",
    "Beds",
    "Baths",
    "Sq Ft",
    "Listed Rent",
    "Required Fees/mo",
    "Promo Raw",
    "Lease Term Mo",
    "Discount Est/mo",
    "Availability Date",
    "Closest Station",
    "Walk Distance Mi",
    "Walk Time Min",
    "Shuttle",
    "Review Score",
    "Review Count",
    "Review Source",
    "Review Summary",
    "Official URL",
    "Listing URL",
    "Transit URL",
    "Review URL",
    "Source Type",
    "Verification Tier",
    "Status",
    "Last Verified At",
    "Notes",
]
CANDIDATE_HEADERS = [
    "Research Date",
    "Researcher",
    "Apartment Name",
    "City",
    "Address",
    "Property Type",
    "Unit/Floorplan",
    "Beds",
    "Baths",
    "Sq Ft",
    "Listed Rent",
    "Required Fees/mo",
    "Promo Raw",
    "Lease Term Mo",
    "Discount Est/mo",
    "Net Monthly Cost",
    "Price/Sq Ft",
    "Availability Date",
    "Closest Station",
    "Walk Distance Mi",
    "Walk Time Min",
    "Shuttle",
    "Review Score",
    "Review Count",
    "Review Source",
    "Review Summary",
    "Official URL",
    "Listing URL",
    "Transit URL",
    "Review URL",
    "Source Type",
    "Verification Tier",
    "Status",
    "Last Verified At",
    "Freshness Hours",
    "Priority Score",
    "Duplicate Flag",
    "Unique Key",
    "Notes",
]
AUDIT_HEADERS = ["Timestamp", "Action", "Sheet", "Property Name", "Unique Key", "Details", "Source URL"]


HEADER_FILL = PatternFill("solid", fgColor="2F5D7E")
SECTION_FILL = PatternFill("solid", fgColor="E8F2FB")
FORMULA_FILL = PatternFill("solid", fgColor="EDF7F2")
CAUTION_FILL = PatternFill("solid", fgColor="FFF4E5")
GRAY_FILL = PatternFill("solid", fgColor="F3F4F6")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def load_data():
    with DATA_PATH.open() as handle:
        return json.load(handle)


def maybe_parse_datetime(value):
    if not isinstance(value, str):
        return value
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return value


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def apply_grid(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = THIN_BORDER
                if ws.title not in {"Controls"}:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_controls(ws, control_rows):
    ws.append(CONTROL_HEADERS)
    for row in control_rows:
        ws.append(row)
    ws["E2"] = "Tracker Notes"
    ws["E2"].font = BOLD_FONT
    ws["E2"].fill = SECTION_FILL
    ws["E3"] = "Use the Controls tab to tune budget and scoring before sorting the Candidates tab."
    ws["E4"] = "Rows older than the stale window should be reviewed before touring or applying."
    ws["E5"] = "Priority Score is a weighted score out of 100 based on cost, transit, reviews, size, and freshness."
    for cell in ("E2", "E3", "E4", "E5"):
        ws[cell].border = THIN_BORDER
        ws[cell].alignment = Alignment(wrap_text=True, vertical="top")
    style_header(ws)
    for row_idx in range(2, ws.max_row + 1):
        ws[f"A{row_idx}"].font = BOLD_FONT
        ws[f"A{row_idx}"].fill = GRAY_FILL
        ws[f"B{row_idx}"].fill = SECTION_FILL
    ws["B15"] = "=TODAY()"
    ws["A15"] = "Today"
    ws["C15"] = "Workbook formula value."
    ws["B16"] = "=NOW()"
    ws["A16"] = "Now"
    ws["C16"] = "Workbook formula value."
    for cell in ("A15", "A16"):
        ws[cell].font = BOLD_FONT
        ws[cell].fill = GRAY_FILL
    for cell in ("B15", "B16"):
        ws[cell].fill = FORMULA_FILL
        ws[cell].number_format = "yyyy-mm-dd hh:mm"
    set_widths(ws, {"A": 24, "B": 16, "C": 62, "E": 22, "F": 64})


def write_records(ws, headers, rows):
    ws.append(headers)
    for record in rows:
        ws.append([maybe_parse_datetime(record.get(header, "")) for header in headers])
    style_header(ws)
    apply_grid(ws)


def add_candidates(ws, records):
    ws.append(CANDIDATE_HEADERS)
    for record in records:
        row = [
            maybe_parse_datetime(record["Research Date"]),
            record["Researcher"],
            record["Apartment Name"],
            record["City"],
            record["Address"],
            record["Property Type"],
            record["Unit/Floorplan"],
            record["Beds"],
            record["Baths"],
            record["Sq Ft"],
            record["Listed Rent"],
            record["Required Fees/mo"],
            record["Promo Raw"],
            record["Lease Term Mo"],
            record["Discount Est/mo"],
            "",
            "",
            maybe_parse_datetime(record["Availability Date"]),
            record["Closest Station"],
            record["Walk Distance Mi"],
            record["Walk Time Min"],
            record["Shuttle"],
            record["Review Score"],
            record["Review Count"],
            record["Review Source"],
            record["Review Summary"],
            record["Official URL"],
            record["Listing URL"],
            record["Transit URL"],
            record["Review URL"],
            record["Source Type"],
            record["Verification Tier"],
            record["Status"],
            maybe_parse_datetime(record["Last Verified At"]),
            "",
            "",
            "",
            "",
            record["Notes"],
        ]
        ws.append(row)

    for row_idx in range(2, ws.max_row + 1):
        ws[f"P{row_idx}"] = f'=IF(K{row_idx}="","",K{row_idx}+IF(L{row_idx}="",0,L{row_idx})-IF(O{row_idx}="",0,O{row_idx}))'
        ws[f"Q{row_idx}"] = f'=IFERROR(P{row_idx}/J{row_idx},"")'
        ws[f"AI{row_idx}"] = f'=IF(AH{row_idx}="","",(NOW()-AH{row_idx})*24)'
        ws[f"AJ{row_idx}"] = (
            f'=IF(C{row_idx}="","",'
            f'IF(AG{row_idx}="REJECTED",-1,'
            f'ROUND(('
            f'MAX(0,MIN(1,(Controls!$B$3-P{row_idx})/MAX(1,Controls!$B$3-Controls!$B$2)))*Controls!$B$6+'
            f'IF(U{row_idx}="",IF(V{row_idx}="Yes",0.6,0),MAX(0,MIN(1,(Controls!$B$4-U{row_idx})/MAX(1,Controls!$B$4))))*Controls!$B$7+'
            f'IF(W{row_idx}="",0,MAX(0,MIN(1,(W{row_idx}-Controls!$B$5)/MAX(0.1,5-Controls!$B$5))))*Controls!$B$8+'
            f'IF(J{row_idx}="",0,MIN(1,J{row_idx}/1100))*Controls!$B$9+'
            f'IF(AI{row_idx}="",0,MAX(0,MIN(1,(Controls!$B$11-AI{row_idx})/MAX(1,Controls!$B$11))))*Controls!$B$10'
            f')/SUM(Controls!$B$6:Controls!$B$10)*100,2)))'
        )
        ws[f"AK{row_idx}"] = f'=IF(AL{row_idx}="","",IF(COUNTIF($AL:$AL,AL{row_idx})>1,"CHECK",""))'
        ws[f"AL{row_idx}"] = f'=IF(OR(C{row_idx}="",J{row_idx}="",K{row_idx}=""),"",LOWER(TRIM(C{row_idx}))&"|"&TEXT(J{row_idx},"0")&"|"&TEXT(K{row_idx},"0.00"))'

    style_header(ws)
    apply_grid(ws)
    formula_cols = ("P", "Q", "AI", "AJ", "AK", "AL")
    caution_cols = ("M", "AM")
    for row_idx in range(2, ws.max_row + 1):
        for col in formula_cols:
            ws[f"{col}{row_idx}"].fill = FORMULA_FILL
        for col in caution_cols:
            ws[f"{col}{row_idx}"].fill = CAUTION_FILL
    set_widths(
        ws,
        {
            "A": 14, "B": 12, "C": 22, "D": 14, "E": 34, "F": 20, "G": 16, "H": 8, "I": 8,
            "J": 10, "K": 12, "L": 14, "M": 32, "N": 12, "O": 14, "P": 14, "Q": 12, "R": 14,
            "S": 16, "T": 12, "U": 12, "V": 10, "W": 11, "X": 12, "Y": 18, "Z": 48, "AA": 24,
            "AB": 24, "AC": 24, "AD": 24, "AE": 15, "AF": 18, "AG": 12, "AH": 18, "AI": 13,
            "AJ": 13, "AK": 12, "AL": 30, "AM": 48
        },
    )


def add_audit_log(ws, rows):
    ws.append(AUDIT_HEADERS)
    for row in rows:
        ws.append([maybe_parse_datetime(row[0]), *row[1:]])
    style_header(ws)
    apply_grid(ws)
    set_widths(ws, {"A": 18, "B": 20, "C": 14, "D": 20, "E": 28, "F": 64, "G": 42})


def format_dates_and_money(wb):
    intake = wb["Intake"]
    for row_idx in range(2, intake.max_row + 1):
        intake[f"A{row_idx}"].number_format = "yyyy-mm-dd"
        if hasattr(intake[f"P{row_idx}"].value, "year"):
            intake[f"P{row_idx}"].number_format = "yyyy-mm-dd"
        for col in ("K", "L", "O"):
            intake[f"{col}{row_idx}"].number_format = '$#,##0.00'
        for col in ("R", "S", "U"):
            intake[f"{col}{row_idx}"].number_format = "0.0"
        intake[f"V{row_idx}"].number_format = "0"

    candidates = wb["Candidates"]
    for row_idx in range(2, candidates.max_row + 1):
        if hasattr(candidates[f"A{row_idx}"].value, "year"):
            candidates[f"A{row_idx}"].number_format = "yyyy-mm-dd"
        if hasattr(candidates[f"R{row_idx}"].value, "year"):
            candidates[f"R{row_idx}"].number_format = "yyyy-mm-dd"
        if hasattr(candidates[f"AH{row_idx}"].value, "year"):
            candidates[f"AH{row_idx}"].number_format = "yyyy-mm-dd hh:mm"
        for col in ("K", "L", "O", "P"):
            candidates[f"{col}{row_idx}"].number_format = '$#,##0.00'
        for col in ("T", "Q", "AI", "AJ", "W"):
            candidates[f"{col}{row_idx}"].number_format = "0.00"
        candidates[f"X{row_idx}"].number_format = "0"
    controls = wb["Controls"]
    controls["B15"].number_format = "yyyy-mm-dd"
    controls["B16"].number_format = "yyyy-mm-dd hh:mm"


def build_workbook():
    data = load_data()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    controls = wb.create_sheet("Controls")
    add_controls(controls, data["controls"])

    seeds = wb.create_sheet("Seed Sources")
    write_records(seeds, SEED_HEADERS, data["seed_sources"])
    set_widths(seeds, {"A": 18, "B": 56, "C": 20, "D": 14, "E": 34, "F": 18, "G": 28, "H": 30, "I": 30, "J": 16})

    queue = wb.create_sheet("Search Queue")
    write_records(queue, QUEUE_HEADERS, data["search_queue"])
    set_widths(queue, {"A": 10, "B": 18, "C": 14, "D": 16, "E": 56, "F": 18, "G": 16, "H": 16, "I": 12, "J": 14, "K": 56})

    intake = wb.create_sheet("Intake")
    write_records(intake, INTAKE_HEADERS, data["intake"])
    set_widths(
        intake,
        {
            "A": 14, "B": 12, "C": 20, "D": 20, "E": 14, "F": 34, "G": 16, "H": 8, "I": 8,
            "J": 10, "K": 12, "L": 14, "M": 32, "N": 12, "O": 14, "P": 14, "Q": 16, "R": 12,
            "S": 12, "T": 10, "U": 11, "V": 12, "W": 18, "X": 48, "Y": 24, "Z": 24, "AA": 24,
            "AB": 24, "AC": 15, "AD": 18, "AE": 12, "AF": 18, "AG": 48
        },
    )

    candidates = wb.create_sheet("Candidates")
    add_candidates(candidates, data["intake"])

    audit = wb.create_sheet("Audit Log")
    add_audit_log(audit, data["audit_log"])

    format_dates_and_money(wb)

    metadata = wb.create_sheet("Workbook Info")
    metadata.append(["Field", "Value"])
    metadata.append(["Generated On", datetime.now().strftime("%Y-%m-%d %H:%M")])
    metadata.append(["Input Data", str(DATA_PATH)])
    metadata.append(["Output File", str(OUTPUT_PATH)])
    metadata.append(["Purpose", "Google-Sheets-ready tracker for South Bay 2B2B apartment search."])
    style_header(metadata)
    apply_grid(metadata)
    set_widths(metadata, {"A": 20, "B": 72})

    wb.save(OUTPUT_PATH)
    print(f"Wrote workbook to {OUTPUT_PATH}")


if __name__ == "__main__":
    build_workbook()
